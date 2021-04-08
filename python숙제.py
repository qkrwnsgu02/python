from openpyxl import load_workbook
from openpyxl import Workbook
from glob import *
import os

Myfiles = [i for i in glob('*.xlsx')]

total_student = []

for item in Myfiles:
    my_workbook = load_workbook(item, data_only = True)
    my_worksheet = my_workbook['Sheet1']
    my_list = []
    my_list.append(my_worksheet['A2'].value)
    my_list.append(my_worksheet['B2'].value)
    my_list.append(my_worksheet['C2'].value)
    my_list.append(my_worksheet['D2'].value)
    total_student.append(my_list)
    print(my_list)



total_student_by_group = {}
for i in range (10):
    total_student_by_group["group"+str(i+1)] = {}
print(total_student_by_group)


assign_dict = {}
for i in range (10):
    assign_dict[i+1] = "group"+str(i+1)
print(assign_dict)

student_number_tracker = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

print(total_student)
# 학생 조에 따라 분류
for item in total_student:
    group = item[2]
    group_name = assign_dict[group]
    total_student_by_group[group_name]["student_"+str(student_number_tracker[group-1]+1)] = item
    student_number_tracker[group-1] += 1

print(total_student_by_group)
print(student_number_tracker)


my_writing_wb = Workbook()


for i in range(10):
    write_ws = my_writing_wb.create_sheet("jo"+str(i+1))


for i in range(10):

    load_ws = my_writing_wb["jo"+str(i+1)]

    load_ws.append(["stu_num", "name", "group_id", "git"])

    tempList = list(total_student_by_group["group"+str(i+1)].values())
    print(tempList)
    for j in range(4):
        try:
            load_ws.append(tempList[j])
        except:
            pass


my_writing_wb.remove(my_writing_wb['Sheet'])

my_writing_wb.save("group_python.xlsx")
